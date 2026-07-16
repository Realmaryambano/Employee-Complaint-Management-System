# ==============================
# Flask Imports
# ==============================
from flask import Flask, render_template, request, redirect, url_for, session, send_file

# ==============================
# Standard Library Imports
# ==============================
import os
import io
import tempfile
from datetime import datetime
from re import search

# ==============================
# Database Imports
# ==============================
from database.db import get_connection

# ==============================
# Excel Export (openpyxl)
# ==============================
from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)

# ==============================
# PDF Export (ReportLab)
# ==============================
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import (
    getSampleStyleSheet,
    ParagraphStyle
)
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.units import inch

from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer
)

# ==============================
# File Handling
# ==============================
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config["UPLOAD_FOLDER"] = "static/uploads"
# ==========================================
# Secret Key
# ==========================================

app.secret_key = "ecms_secret_key_2026"


# ==========================================
# Login
# ==========================================

@app.route("/", methods=["GET", "POST"])
def login():

    error = None

    if request.method == "POST":

        username = request.form["username"]
        password = request.form["password"]

        connection = get_connection()
        cursor = connection.cursor()

        cursor.execute("""
            SELECT Employee_ID, Role
            FROM Employees
            WHERE Email = :email
            AND Password = :password
        """,
        email=username,
        password=password)

        user = cursor.fetchone()

        cursor.close()
        connection.close()

        if user:

            session["employee_id"] = user[0]
            session["role"] = user[1]

            if user[1] == "Admin":
                return redirect(url_for("it_dashboard"))

            return redirect(url_for("employee_dashboard"))

        error = "Invalid Email or Password."

    return render_template(
        "login.html",
        error=error
    )


# ==========================================
# Employee Dashboard
# ==========================================
@app.route("/employee-dashboard")
def employee_dashboard():

    employee_id = session.get("employee_id")

    if not employee_id:
        return redirect(url_for("login"))

    connection = get_connection()
    cursor = connection.cursor()

    cursor.execute("""
        SELECT
            e.Employee_Name,
            e.Employee_ID,
            d.Department_Name,
            e.Email

        FROM Employees e

        JOIN Departments d
        ON e.Department_ID = d.Department_ID

        WHERE e.Employee_ID = :employee_id
    """,
    employee_id=employee_id)

    employee = cursor.fetchone()


    # ==============================
    # Employee Complaint Statistics
    # ==============================

    cursor.execute("""
        SELECT COUNT(*)
        FROM Complaints
        WHERE Employee_ID = :employee_id
    """,
    employee_id=employee_id)

    total = cursor.fetchone()[0]


    cursor.execute("""
        SELECT COUNT(*)
        FROM Complaints
        WHERE Employee_ID = :employee_id
        AND Status = 'Pending'
    """,
    employee_id=employee_id)

    pending = cursor.fetchone()[0]


    cursor.execute("""
        SELECT COUNT(*)
        FROM Complaints
        WHERE Employee_ID = :employee_id
        AND Status = 'In Progress'
    """,
    employee_id=employee_id)

    progress = cursor.fetchone()[0]


    cursor.execute("""
        SELECT COUNT(*)
        FROM Complaints
        WHERE Employee_ID = :employee_id
        AND Status = 'Resolved'
    """,
    employee_id=employee_id)

    resolved = cursor.fetchone()[0]


    cursor.close()
    connection.close()


    return render_template(
        "employee_dashboard.html",
        employee=employee,
        total=total,
        pending=pending,
        progress=progress,
        resolved=resolved
    )

# ==========================================
# Complaint Form
# ==========================================

@app.route("/complaint-form", methods=["GET", "POST"])
def complaint_form():

    employee_id = session.get("employee_id")

    if not employee_id:
        return redirect(url_for("login"))

    connection = get_connection()
    cursor = connection.cursor()

    if request.method == "POST":

        issue_title = request.form["issue_title"]
        category = request.form["category"]
        priority = request.form["priority"]
        store_location = request.form["store_location"]
        contact_number = request.form["contact_number"]
        device_name = request.form["device_name"]
        description = request.form["description"]
        attachment = request.files.get("attachment")

        filename = None

        if attachment and attachment.filename != "":
            filename = secure_filename(attachment.filename)
            attachment.save(
                os.path.join(app.config["UPLOAD_FOLDER"], filename)
            )

        cursor.execute("""
            INSERT INTO Complaints
            (
                Employee_ID,
                Issue_Title,
                Category,
                Priority,
                Store_Location,
                Contact_Number,
                Device_Name,
                Description,
                Attachment
            )

            VALUES
            (
                :employee_id,
                :issue_title,
                :category,
                :priority,
                :store_location,
                :contact_number,
                :device_name,
                :description,
                :attachment
            )
        """,
    employee_id=employee_id,
    issue_title=issue_title,
    category=category,
    priority=priority,
    store_location=store_location,
    contact_number=contact_number,
    device_name=device_name,
    description=description,
    attachment=filename
    )

        connection.commit()

        cursor.execute("""
            SELECT MAX(Complaint_ID)
            FROM Complaints
            WHERE Employee_ID=:employee_id
        """,
        employee_id=employee_id)

        complaint_id = cursor.fetchone()[0]

        cursor.close()
        connection.close()

        return render_template(
            "complaint_success.html",
            complaint_id=complaint_id
        )

    cursor.execute("""
        SELECT
            e.Employee_Name,
            e.Employee_ID,
            d.Department_Name,
            e.Email,
            TO_CHAR(SYSDATE,'FMDD Month YYYY'),
            TO_CHAR(SYSDATE,'HH:MI AM')
        FROM Employees e

        JOIN Departments d
        ON e.Department_ID = d.Department_ID

        WHERE e.Employee_ID = :employee_id
    """,
    employee_id=employee_id)

    employee = cursor.fetchone()

    cursor.close()
    connection.close()

    return render_template(
        "complaint_form.html",
        employee=employee
    )


# ==========================================
# My Complaints
# ==========================================
@app.route("/my-complaints")
def my_complaints():

    employee_id = session.get("employee_id")

    if not employee_id:
        return redirect(url_for("login"))

    connection = get_connection()
    cursor = connection.cursor()

    search = request.args.get("search", "")
    status_filter = request.args.get("status", "")


    query = """
        SELECT
            Complaint_ID,
            Issue_Title,
            Category,
            Priority,
            Date_Submitted,
            Status,
            Remarks

        FROM Complaints

        WHERE Employee_ID = :employee_id
    """


    params = {
        "employee_id": employee_id
    }


    if search:

        query += """
        AND LOWER(Issue_Title) LIKE LOWER(:search)
        """

        params["search"] = f"%{search}%"



    if status_filter:

        query += """
        AND Status = :status
        """

        params["status"] = status_filter



    query += """
        ORDER BY Date_Submitted DESC
    """


    cursor.execute(query, params)


    complaints = []


    for row in cursor.fetchall():

        row = list(row)

        if row[6]:

            try:
                row[6] = row[6].read()

            except:
                pass


        complaints.append(tuple(row))


    cursor.close()
    connection.close()


    return render_template(
        "my_complaints.html",
        complaints=complaints,
        search=search,
        status_filter=status_filter
    )


@app.route("/it-dashboard")
def it_dashboard():

    connection = get_connection()
    cursor = connection.cursor()

    search = request.args.get("search", "")
    status_filter = request.args.get("status", "")

    # Define filter_sql to prevent NameError
    filter_sql = ""

    # Dashboard Counts
    cursor.execute(f"""
    SELECT COUNT(*)
    FROM Complaints c
    {filter_sql}
    """)
    total = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM COMPLAINTS WHERE STATUS='Pending'")
    pending = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM COMPLAINTS WHERE STATUS='In Progress'")
    progress = cursor.fetchone()[0]

    cursor.execute(f"""
    SELECT COUNT(*)
    FROM Complaints c
    {filter_sql}
    {"AND" if filter_sql else "WHERE"} Status='Resolved'
    """)
    resolved = cursor.fetchone()[0]

    # Complaint Table
    query = """
        SELECT
            COMPLAINT_ID,
            EMPLOYEE_ID,
            CATEGORY,
            ISSUE_TITLE,
            PRIORITY,
            STATUS
        FROM COMPLAINTS
        WHERE 1=1
    """

    params = {}

    if search:
        search_value = search.upper().replace("ECMS-", "")
        query += """
            AND (
                TO_CHAR(COMPLAINT_ID) LIKE :search
                OR LOWER(ISSUE_TITLE) LIKE LOWER(:search)
            )
        """
        params["search"] = f"%{search_value}%"

    if status_filter:
        query += """
            AND STATUS = :status
        """
        params["status"] = status_filter

    query += """
        ORDER BY COMPLAINT_ID DESC
    """

    cursor.execute(query, params)

    complaints = cursor.fetchall()
    cursor.close()
    connection.close()

    return render_template(
        "it_dashboard.html",
        complaints=complaints,
        total=total,
        pending=pending,
        progress=progress,
        resolved=resolved,
        search=search,
        status_filter=status_filter
    )
# ==========================================
# Reports & Analytics
# ==========================================

@app.route("/reports")
def reports():

    if session.get("role") != "Admin":
        return redirect(url_for("login"))

    period = request.args.get("period", "all")
    print("Selected Period:", period)

    connection = get_connection()
    cursor = connection.cursor()

    # ----------------------------
    # Date Filter
    # ----------------------------

    filter_sql = ""

    if period == "today":

        filter_sql = """
        WHERE TRUNC(c.Date_Submitted) = TRUNC(SYSDATE)
        """

    elif period == "yesterday":

        filter_sql = """
        WHERE TRUNC(c.Date_Submitted) = TRUNC(SYSDATE)-1
        """

    elif period == "week":

        filter_sql = """
        WHERE c.Date_Submitted >= SYSDATE-7
        """

    elif period == "month":

        filter_sql = """
        WHERE EXTRACT(MONTH FROM c.Date_Submitted)=EXTRACT(MONTH FROM SYSDATE)
        AND EXTRACT(YEAR FROM c.Date_Submitted)=EXTRACT(YEAR FROM SYSDATE)
        """

    elif period == "lastmonth":

        filter_sql = """
        WHERE c.Date_Submitted >= ADD_MONTHS(TRUNC(SYSDATE,'MM'),-1)
        AND c.Date_Submitted < TRUNC(SYSDATE,'MM')
        """

    elif period == "year":

        filter_sql = """
        WHERE EXTRACT(YEAR FROM c.Date_Submitted)=EXTRACT(YEAR FROM SYSDATE)
        """

    elif period == "lastyear":

        filter_sql = """
        WHERE EXTRACT(YEAR FROM c.Date_Submitted)=EXTRACT(YEAR FROM SYSDATE)-1
        """
    
    print(filter_sql)


    # ----------------------------
    # Summary Cards
    # ----------------------------

    summary_query = f"""

        SELECT

            COUNT(*) AS TOTAL,

            SUM(CASE
                    WHEN Status='Resolved'
                    THEN 1
                    ELSE 0
                END) AS RESOLVED,

            SUM(CASE
                    WHEN Status='Pending'
                    THEN 1
                    ELSE 0
                END) AS PENDING,

            SUM(CASE
                    WHEN Status='In Progress'
                    THEN 1
                    ELSE 0
                END) AS PROGRESS

        FROM Complaints c

        {filter_sql}

    """

    cursor.execute(summary_query)

    row = cursor.fetchone()

    total = row[0] if row[0] else 0
    resolved = row[1] if row[1] else 0
    pending = row[2] if row[2] else 0
    progress = row[3] if row[3] else 0

    # ----------------------------
    # Complaint Details
    # ----------------------------

    query = f"""

        SELECT

            c.Complaint_ID,
            e.Employee_Name,
            d.Department_Name,
            c.Category,
            c.Priority,
            c.Issue_Title,
            c.Status,
            c.Date_Submitted

        FROM Complaints c

        JOIN Employees e
        ON c.Employee_ID = e.Employee_ID

        JOIN Departments d
        ON e.Department_ID = d.Department_ID

        {filter_sql}

        ORDER BY c.Date_Submitted DESC

    """

    cursor.execute(query)

    complaints = cursor.fetchall()

    cursor.close()
    connection.close()

    return render_template(

        "reports.html",

        complaints=complaints,

        total=total,
        resolved=resolved,
        pending=pending,
        progress=progress,

        period=period

    )
# ==========================================
# Export PDF Report
# ==========================================

@app.route("/export-pdf")
def export_pdf():

    if session.get("role") != "Admin":
        return redirect(url_for("login"))

    period = request.args.get("period", "all")

    connection = get_connection()
    cursor = connection.cursor()

    # ==========================================
    # Report Period
    # ==========================================

    filter_sql = ""
    period_name = "All Time"

    if period == "today":
        period_name = "Today"
        filter_sql = "WHERE TRUNC(c.Date_Submitted)=TRUNC(SYSDATE)"

    elif period == "yesterday":
        period_name = "Yesterday"
        filter_sql = "WHERE TRUNC(c.Date_Submitted)=TRUNC(SYSDATE)-1"

    elif period == "week":
        period_name = "Last 7 Days"
        filter_sql = "WHERE c.Date_Submitted>=SYSDATE-7"

    elif period == "month":
        period_name = "This Month"
        filter_sql = """
        WHERE EXTRACT(MONTH FROM c.Date_Submitted)=EXTRACT(MONTH FROM SYSDATE)
        AND EXTRACT(YEAR FROM c.Date_Submitted)=EXTRACT(YEAR FROM SYSDATE)
        """

    elif period == "lastmonth":
        period_name = "Last Month"
        filter_sql = """
        WHERE c.Date_Submitted>=ADD_MONTHS(TRUNC(SYSDATE,'MM'),-1)
        AND c.Date_Submitted<TRUNC(SYSDATE,'MM')
        """

    elif period == "year":
        period_name = "This Year"
        filter_sql = """
        WHERE EXTRACT(YEAR FROM c.Date_Submitted)=EXTRACT(YEAR FROM SYSDATE)
        """

    elif period == "lastyear":
        period_name = "Last Year"
        filter_sql = """
        WHERE EXTRACT(YEAR FROM c.Date_Submitted)=EXTRACT(YEAR FROM SYSDATE)-1
        """

    # ==========================================
    # Summary
    # ==========================================

    summary_query = f"""
    SELECT
        COUNT(*),
        SUM(CASE WHEN Status='Resolved' THEN 1 ELSE 0 END),
        SUM(CASE WHEN Status='Pending' THEN 1 ELSE 0 END),
        SUM(CASE WHEN Status='In Progress' THEN 1 ELSE 0 END)
    FROM Complaints c
    {filter_sql}
    """

    cursor.execute(summary_query)

    summary = cursor.fetchone()

    total = summary[0] or 0
    resolved = summary[1] or 0
    pending = summary[2] or 0
    progress = summary[3] or 0

    # ==========================================
    # Complaint Data
    # ==========================================

    query = f"""
    SELECT

        c.Complaint_ID,
        e.Employee_Name,
        d.Department_Name,
        c.Category,
        c.Priority,
        c.Issue_Title,
        c.Status,
        TO_CHAR(c.Date_Submitted,'DD-MON-YYYY')

    FROM Complaints c

    JOIN Employees e
        ON c.Employee_ID=e.Employee_ID

    JOIN Departments d
        ON e.Department_ID=d.Department_ID

    {filter_sql}

    ORDER BY c.Date_Submitted DESC
    """

    cursor.execute(query)

    complaints = cursor.fetchall()

    cursor.close()
    connection.close()

    # ==========================================
    # PDF
    # ==========================================

    temp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")

    doc = SimpleDocTemplate(
        temp.name,
        pagesize=landscape(letter),
        rightMargin=25,
        leftMargin=25,
        topMargin=25,
        bottomMargin=25
    )

    styles = getSampleStyleSheet()

    title = ParagraphStyle(
        "Title",
        parent=styles["Heading1"],
        alignment=1,
        fontSize=22,
        textColor=colors.HexColor("#0D47A1"),
        spaceAfter=8
    )

    heading = ParagraphStyle(
        "Heading",
        parent=styles["Heading2"],
        alignment=1,
        textColor=colors.HexColor("#1565C0"),
        spaceAfter=15
    )

    normal = styles["Normal"]

    elements = []

    # ==========================================
    # Header
    # ==========================================

    elements.append(Paragraph("<b>BONANZA SATRANGI</b>", title))

    elements.append(
        Paragraph(
            "Employee Complaint Management System",
            heading
        )
    )

    from datetime import datetime

    elements.append(
        Paragraph(
            f"<b>Report Period:</b> {period_name}",
            normal
        )
    )

    elements.append(
        Paragraph(
            f"<b>Generated On:</b> {datetime.now().strftime('%d %B %Y, %I:%M %p')}",
            normal
        )
    )

    elements.append(Spacer(1, 15))

    # ==========================================
    # Summary Table
    # ==========================================

    summary_data = [

        ["Total", "Resolved", "Pending", "In Progress"],

        [str(total), str(resolved), str(pending), str(progress)]

    ]

    summary_table = Table(summary_data, colWidths=[120]*4)

    summary_table.setStyle(TableStyle([

        ('BACKGROUND',(0,0),(-1,0),colors.HexColor("#0D47A1")),
        ('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTNAME',(0,0),(-1,0),"Helvetica-Bold"),
        ('FONTSIZE',(0,0),(-1,0),11),

        ('BACKGROUND',(0,1),(-1,1),colors.HexColor("#F4F7FB")),

        ('GRID',(0,0),(-1,-1),0.6,colors.grey),

        ('ALIGN',(0,0),(-1,-1),"CENTER"),

        ('BOTTOMPADDING',(0,0),(-1,0),10),

        ('TOPPADDING',(0,1),(-1,1),10)

    ]))

    elements.append(summary_table)

    elements.append(Spacer(1,20))

    # ==========================================
    # Complaint Table
    # ==========================================

    data = [[

        "ID",
        "Employee",
        "Department",
        "Category",
        "Priority",
        "Issue",
        "Status",
        "Date"

    ]]

    for row in complaints:

        data.append([

            f"ECMS-{row[0]}",
            row[1],
            row[2],
            row[3],
            row[4],
            row[5],
            row[6],
            row[7]

        ])

    table = Table(
        data,
        colWidths=[60,110,90,80,70,180,80,80]
    )

    table.setStyle(TableStyle([

        ('BACKGROUND',(0,0),(-1,0),colors.HexColor("#0D47A1")),
        ('TEXTCOLOR',(0,0),(-1,0),colors.white),

        ('FONTNAME',(0,0),(-1,0),"Helvetica-Bold"),

        ('FONTSIZE',(0,0),(-1,0),10),

        ('BACKGROUND',(0,1),(-1,-1),colors.white),

        ('GRID',(0,0),(-1,-1),0.5,colors.grey),

        ('VALIGN',(0,0),(-1,-1),"MIDDLE"),

        ('ALIGN',(0,0),(-1,-1),"CENTER"),

        ('BOTTOMPADDING',(0,0),(-1,0),10),

        ('TOPPADDING',(0,1),(-1,-1),7)

    ]))

    elements.append(table)

    elements.append(Spacer(1,20))

    elements.append(
        Paragraph(
            "<i>Generated by Employee Complaint Management System (ECMS)</i>",
            styles["Italic"]
        )
    )

    # ==========================================
    # Build PDF
    # ==========================================

    doc.build(elements)

    return send_file(

        temp.name,

        as_attachment=True,

        download_name="ECMS_Complaint_Report.pdf",

        mimetype="application/pdf"

    )
# ==========================================
# Export Excel Report
# ==========================================


@app.route("/export-excel/<period>")
def export_excel(period):

    if session.get("role") != "Admin":
        return redirect(url_for("login"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Complaint Report"

    # ==========================
    # Styles
    # ==========================

    blue = PatternFill("solid", fgColor="0D47A1")
    light_blue = PatternFill("solid", fgColor="1976D2")
    grey = PatternFill("solid", fgColor="F3F4F6")
    green = PatternFill("solid", fgColor="4CAF50")
    orange = PatternFill("solid", fgColor="FB8C00")
    sky = PatternFill("solid", fgColor="2196F3")

    white_font = Font(color="FFFFFF", bold=True)
    bold = Font(bold=True)
    title = Font(size=18, bold=True)
    subtitle = Font(size=14, bold=True)

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ==========================
    # Header
    # ==========================

    ws.merge_cells("A1:H1")
    ws["A1"] = "BONANZA SATRANGI"
    ws["A1"].fill = blue
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    ws["A1"].alignment = center

    ws.merge_cells("A2:H2")
    ws["A2"] = "Employee Complaint Management System"
    ws["A2"].fill = light_blue
    ws["A2"].font = white_font
    ws["A2"].alignment = center

    ws.merge_cells("A4:H4")
    ws["A4"] = "EMPLOYEE COMPLAINT ANALYTICS REPORT"
    ws["A4"].font = title
    ws["A4"].alignment = center

    # ==========================
    # Database Connection
    # ==========================

    connection = get_connection()
    cursor = connection.cursor()

    # ==========================
    # Date Filter
    # ==========================

    filter_sql = ""
    period_name = "All Time"

    if period == "today":
        period_name = "Today"
        filter_sql = """
        WHERE TRUNC(c.Date_Submitted)=TRUNC(SYSDATE)
        """

    elif period == "yesterday":
        period_name = "Yesterday"
        filter_sql = """
        WHERE TRUNC(c.Date_Submitted)=TRUNC(SYSDATE)-1
        """

    elif period == "week":
        period_name = "Last 7 Days"
        filter_sql = """
        WHERE c.Date_Submitted>=SYSDATE-7
        """

    elif period == "month":
        period_name = "This Month"
        filter_sql = """
        WHERE EXTRACT(MONTH FROM c.Date_Submitted)=EXTRACT(MONTH FROM SYSDATE)
        AND EXTRACT(YEAR FROM c.Date_Submitted)=EXTRACT(YEAR FROM SYSDATE)
        """

    elif period == "lastmonth":
        period_name = "Last Month"
        filter_sql = """
        WHERE c.Date_Submitted>=ADD_MONTHS(TRUNC(SYSDATE,'MM'),-1)
        AND c.Date_Submitted<TRUNC(SYSDATE,'MM')
        """

    elif period == "year":
        period_name = "This Year"
        filter_sql = """
        WHERE EXTRACT(YEAR FROM c.Date_Submitted)=EXTRACT(YEAR FROM SYSDATE)
        """

    elif period == "lastyear":
        period_name = "Last Year"
        filter_sql = """
        WHERE EXTRACT(YEAR FROM c.Date_Submitted)=EXTRACT(YEAR FROM SYSDATE)-1
        """

    # ==========================
    # Report Metadata (Fixed placement to prevent Crash)
    # ==========================


    ws["A6"] = "Generated On"
    ws["A6"].font = bold
    ws["B6"] = datetime.now().strftime("%d-%b-%Y %I:%M %p")

    ws["D6"] = "Report Period"
    ws["D6"].font = bold
    ws["E6"] = period_name

    # ==========================
    # Summary (Optimized with dynamic Period Filter)
    # ==========================

    summary_query = f"""
    SELECT 
        COUNT(*) AS TOTAL,
        SUM(CASE WHEN Status='Resolved' THEN 1 ELSE 0 END),
        SUM(CASE WHEN Status='Pending' THEN 1 ELSE 0 END),
        SUM(CASE WHEN Status='In Progress' THEN 1 ELSE 0 END)
    FROM Complaints c
    {filter_sql}
    """

    cursor.execute(summary_query)
    row_data = cursor.fetchone()

    total = row_data[0] or 0
    resolved = row_data[1] or 0
    pending = row_data[2] or 0
    progress = row_data[3] or 0

    ws.merge_cells("A8:H8")
    ws["A8"] = "SUMMARY"
    ws["A8"].fill = blue
    ws["A8"].font = white_font
    ws["A8"].alignment = center

    summary = [
        ("Total Complaints", total),
        ("Resolved", resolved),
        ("Pending", pending),
        ("In Progress", progress)
    ]

    row = 9
    for item, value in summary:
        ws[f"A{row}"] = item
        ws[f"A{row}"].font = bold
        ws[f"A{row}"].fill = grey

        ws[f"B{row}"] = value
        ws[f"B{row}"].alignment = center

        ws[f"A{row}"].border = border
        ws[f"B{row}"].border = border
        row += 1

    # ==========================
    # Table Title
    # ==========================

    start = row + 2
    ws.merge_cells(f"A{start}:H{start}")
    ws[f"A{start}"] = "COMPLAINT DETAILS"
    ws[f"A{start}"].fill = blue
    ws[f"A{start}"].font = white_font
    ws[f"A{start}"].alignment = center

    start += 1
    headers = [
        "Complaint ID",
        "Employee",
        "Department",
        "Category",
        "Priority",
        "Issue Title",
        "Status",
        "Date Submitted"
    ]

    for col, head in enumerate(headers, 1):
        cell = ws.cell(row=start, column=col)
        cell.value = head
        cell.fill = blue
        cell.font = white_font
        cell.alignment = center
        cell.border = border

    # ==========================
    # Complaint Data Table
    # ==========================

    query = f"""
    SELECT 
        c.Complaint_ID,
        e.Employee_Name,
        d.Department_Name,
        c.Category,
        c.Priority,
        c.Issue_Title,
        c.Status,
        c.Date_Submitted
    FROM Complaints c
    JOIN Employees e ON c.Employee_ID=e.Employee_ID
    JOIN Departments d ON e.Department_ID=d.Department_ID
    {filter_sql}
    ORDER BY c.Date_Submitted DESC
    """

    cursor.execute(query)
    r = start + 1

    for complaint in cursor.fetchall():
        values = [
            f"ECMS-{complaint[0]}",
            complaint[1],
            complaint[2],
            complaint[3],
            complaint[4],
            complaint[5],
            complaint[6],
            str(complaint[7])
        ]

        for c, value in enumerate(values, 1):
            cell = ws.cell(row=r, column=c)
            cell.value = value
            cell.border = border

            if c == 7:
                cell.alignment = center
                if value == "Resolved":
                    cell.fill = green
                elif value == "Pending":
                    cell.fill = orange
                elif value == "In Progress":
                    cell.fill = sky
            elif c in [1, 4, 7, 8]:
                cell.alignment = center
            else:
                cell.alignment = left
        r += 1

    cursor.close()
    connection.close()

    # ==========================
    # Footer
    # ==========================

    r += 2
    ws.merge_cells(f"A{r}:H{r}")
    ws[f"A{r}"] = "Generated by Employee Complaint Management System (ECMS)"
    ws[f"A{r}"].font = Font(italic=True)

    # ==========================
    # Widths
    # ==========================

    widths = {
        "A": 15,
        "B": 25,
        "C": 20,
        "D": 20,
        "E": 15,
        "F": 35,
        "G": 18,
        "H": 24
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # ==========================
    # Download File Generation
    # ==========================

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name="ECMS_Complaint_Report.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
# ==========================================
# Complaint Details
# ==========================================

@app.route("/complaint-details/<int:complaint_id>", methods=["GET", "POST"])
def complaint_details(complaint_id):

    if session.get("role") != "Admin":
        return redirect(url_for("login"))

    connection = get_connection()
    cursor = connection.cursor()

    # --------------------------------------
    # Update Complaint
    # --------------------------------------

    if request.method == "POST":

        status = request.form["status"]
        remarks = request.form["remarks"]

        cursor.execute("""
            UPDATE Complaints

            SET
                Status = :status,
                Remarks = :remarks

            WHERE Complaint_ID = :complaint_id
        """,
        status=status,
        remarks=remarks,
        complaint_id=complaint_id)

        connection.commit()

        cursor.close()
        connection.close()

        return redirect(url_for("it_dashboard"))

    # --------------------------------------
    # Load Complaint Details
    # --------------------------------------

    cursor.execute("""
        SELECT

        c.Complaint_ID,
        e.Employee_Name,
        e.Employee_ID,
        d.Department_Name,
        c.Category,
        c.Priority,
        c.Issue_Title,
        c.Description,
        c.Status,
        c.Remarks,
        c.Date_Submitted,
        c.Attachment

        FROM Complaints c

        JOIN Employees e
        ON c.Employee_ID = e.Employee_ID

        JOIN Departments d
        ON e.Department_ID = d.Department_ID

        WHERE c.Complaint_ID = :complaint_id
    """,
    complaint_id=complaint_id)

    complaint = cursor.fetchone()

    if complaint:

        complaint = list(complaint)


    # Convert CLOB fields before closing connection

    if complaint[7]:
        try:
            complaint[7] = complaint[7].read()
        except:
            pass


    if complaint[9]:
        try:
            complaint[9] = complaint[9].read()
        except:
            pass


    complaint = tuple(complaint)

    cursor.close()
    connection.close()

    return render_template(
        "complaint_details.html",
        complaint=complaint
    )


# ==========================================
# IT Support
# ==========================================

@app.route("/it-support")
def it_support():

    return render_template(
        "it_support.html"
    )


# ==========================================
# Logout
# ==========================================

@app.route("/logout")
def logout():

    session.clear()

    return redirect(
        url_for("login")
    )


# ==========================================
# 404 Error
# ==========================================

@app.errorhandler(404)
def page_not_found(error):

    return """

    <h1>404 - Page Not Found</h1>

    <a href="/">Return Home</a>

    """, 404


# ==========================================
# Oracle Connection Test
# ==========================================

try:

    connection = get_connection()

    print("✅ Connected to Oracle Database Successfully!")

    connection.close()

except Exception as e:

    print("❌ Connection Failed")
    print(e)


# ==========================================
# Run Application
# ==========================================

if __name__ == "__main__":

    app.run(debug=True)
    